import openpyxl
wb=openpyxl.load_workbook("C://data/marks.xlsx")
sheet=wb.active
data=sheet['A3'].value
data3=sheet.cell(row=2,column=3).value
sheet['A3'].value="hello"
print(data)
print(data3)
print(sheet.max_row)
print(sheet.max_column)
for i in range(1,7):
    print(sheet.cell(row=i,colum=1).value)